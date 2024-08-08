import inspect
from PyQt5.QtWidgets import *
from PyQt5.QtCore import Qt, QThread, pyqtSlot, pyqtSignal
import logging
from datetime import datetime
import os
import sys
import time
from openpyxl import load_workbook
from main_ui import Ui_Dialog as Main_Ui
from module.trans import *
from module.form import *
from scapy.all import *
from scapy.layers.inet import UDP
from scapy.layers.inet import IP
from scapy.layers.inet import ICMP
from scapy.layers.inet import TCP

MESSAGE_TYPES={
        "0x00":"REQUEST",
        "0x01":"REQUEST_NO_RET",
        "0x02":"NOTIFICATION",
        "0x40":"REQUEST_ACK",
        "0x41":"REQUEST_NORET_ACK",
        "0x42":"NOTIFICATION_ACK",
        "0x80":"RESPONSE",
        "0x81":"ERROR",
        "0xc0":"RESPONSE_ACK",
        "0xc1":"ERROR_ACK",
        "0x20":"TP_REQUEST",
        "0x21":"TP_REQUEST_NO_RET",
        "0x22":"TP_NOTIFICATION",
        "0xa0":"TP_RESPONSE",
        "0xa1":"TP_ERROR"
    }

def convert_to_hex(code, prefix=True):
    return '0x' + code.hex() if prefix else code.hex()

def parse_ip(summary):
    from_ip, to_ip = summary.split(' > ')
    from_ip = re.findall(r'[0-9]+(?:\.[0-9]+){3}', from_ip)[0]
    to_ip = re.findall(r'[0-9]+(?:\.[0-9]+){3}', to_ip)[0]
    # print(from_ip)
    # print(to_ip)
    return {'from_ip': from_ip, 'to_ip': to_ip}

def response_parser(data):
    # b'E\x00\x000\x1e\xe6\x00\x00@\x11\xb7\x1fd@n\x0bd@n,\xc7E\xc7E\x00\x1cFO
    # \x00\x13                      # Service ID
    # \x05\x01                      # Method ID
    # \x00\x00\x00\x0c              # Length 12
    # \x00\x01                      # Client ID
    # \x00\x01                      # Session ID
    # \x01                          # SOME/IP Version
    # \x01                          # Interface Version
    # \x80                          # Message Type
    # \x00                          # Return Code
    # \x00\x00\x00\x00              # Payload
    # \x00\x00'
    # print(data[0:28])
    data = data[28:]

    service_id = convert_to_hex(data[0:2])
    # print(service_id)

    method_id = convert_to_hex(data[2:4])
    # print(method_id)

    client_id = convert_to_hex(data[8:10])
    # print(client_id)

    session_id = convert_to_hex(data[10:12])
    # print(session_id)

    message_type = convert_to_hex(data[14:15])
    message_type = MESSAGE_TYPES[message_type]
    # print(message_type)

    payload = convert_to_hex(data[16:20], False)
    # print(payload)

    return {'service_id': service_id,
            'method_id': method_id,
            # 'client_id': client_id,
            # 'session_id': session_id,
            'message_type': message_type,
            'payload': payload}

class MainDialog(QDialog):
    def __init__(self):
        # Display minimize, close button
        super().__init__()

        self.setWindowFlag(Qt.WindowMinimizeButtonHint, True)
        self.setWindowFlag(Qt.WindowMaximizeButtonHint, False)
        self.setWindowFlag(Qt.WindowCloseButtonHint, True)

        self.main_ui = Main_Ui()
        self.main_ui.setupUi(self)
        self.setWindowTitle('SOMEIP (240806)')

        self.init_setup()

        self.main_ui.btn_exit_program.clicked.connect(self.exit_program)
        self.main_ui.btn_import_excel.clicked.connect(self.open_import_excel)
        self.main_ui.btn_run_automation.clicked.connect(self.run_automation)
        self.main_ui.btn_send_manual.clicked.connect(self.send_manual)
        self.main_ui.btn_stop_automation.clicked.connect(self.stop_automation)
        # self.main_ui.btn_send_manual_2.clicked.connect(self.send_manual2)
        self.main_ui.btn_clear_log.clicked.connect(self.clean_log)

        self.manual_response_data_list = []


    def clean_log(self):
        self.main_ui.tb_status.clear()

    def send_manual(self):
        try:
            address_sim = self.main_ui.le_simul_ip.text()       # String
            port_sim = self.main_ui.sb_simul_port.value()       # Int
            address_dest = self.main_ui.le_dest_ip.text()       # String
            port_dest = self.main_ui.sb_dest_port.value()       # Int
            protocol_ver = self.main_ui.sb_protocol_ver.value() # Int
            message_id = self.main_ui.le_msg_id.text()          # String
            service_id, method_id = getIds(message_id)
            msg_type = self.main_ui.cb_msg_type.currentText()   # String

            msg_status = self.main_ui.cb_msg_status.currentText()   # String
            client_id = self.main_ui.sp_client_id.value()       # Int
            session_id = self.main_ui.sp_session_id.value()     # Int
            payload = ConvertPayloadToHexString(self.main_ui.le_payload.text())
            payload = hexStringToByte(payload)

            form = SomeIpForm(
                proto=protocol_ver,
                msg_type=msg_type,
                ret_code=msg_status,
                srv_id=service_id,
                method_id=method_id,
                client_id=client_id,
                session_id=session_id,
                payload=payload
            )

            form.setIp(address_sim, address_dest)
            form.setUdp(port_sim, port_dest)
            form.setPacket()
            # r, u = form.sr()

            # summary = r[0][1].summary()
            # original = r[0][1].original
            original = b'E\x00\x000\x1e\xe6\x00\x00@\x11\xb7\x1fd@n\x0bd@n,\xc7E\xc7E\x00\x1cFO\x00\x13\x05\x01\x00\x00\x00\x0c\x00\x01\x00\x01\x01\x01\x80\x00\x00\x00\x00\x00\x00\x00'
            summary = 'IP / UDP 100.64.110.11:51013 > 100.64.110.44:51013 / Raw / Padding'

            response = response_parser(original)
            ip_info = parse_ip(summary)
            response.update(ip_info)

            self.manual_response_data_list.append(response)

            # response = {'from_ip':'100.64.111.44',
            #             'to_ip':'100.64.111.11',
            #             'service_id':'0x0013',
            #             'method_id':'0x0051',
            #             'message_type':'RESPONSE',
            #             'payload':'00000000'}

            log_message = f'[{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}]\n'
            log_message += f'Source IP: {response["from_ip"]} / Dest IP: {response["to_ip"]}\n'
            log_message += f'Service ID: {response["service_id"]}\n'
            log_message += f'Method ID: {response["method_id"]}\n'
            log_message += f'Message Type: {response["message_type"]}\n'
            log_message += f'Payload: {response["payload"]}\n'
            self.main_ui.tb_status.append(log_message)

            # QMessageBox.information(self, "Information", "Transfer Complete.")
        except Exception as e:
            QMessageBox.critical(self, 'Exception', str(e))

    def open_import_excel(self):
        self.main_ui.le_excelfile_path.setText('')
        current_dir = os.path.curdir
        file_name = QFileDialog.getOpenFileName(self, 'Open Excel File', current_dir, 'Excel(*.xlsx)')
        if file_name[0] != '':
            if file_name[0].split('.')[-1] != 'xlsx':
                QMessageBox.warning(self, "Warning", "Please select only excel file!")
                return
            self.main_ui.le_excelfile_path.setText(file_name[0])
            self.main_ui.btn_run_automation.setEnabled(True)
        else:
            self.main_ui.btn_run_automation.setEnabled(False)

    def run_automation(self):
        excel_file_path = self.main_ui.le_excelfile_path.text()

        self.thread = TransferThread(excel_file_path, self.main_ui.sp_interval.value())
        self.thread.countSignal.connect(self.prog_count)
        self.thread.completeSignal.connect(self.thread_complete)
        self.thread.logSignal.connect(self.thread_log)
        self.thread.start()

        self.main_ui.btn_stop_automation.setEnabled(True)

    def stop_automation(self):
        if self.thread is not None:
            self.thread.stop_thread()
            self.main_ui.btn_run_automation.setEnabled(True)
            self.main_ui.btn_stop_automation.setEnabled(False)

    @pyqtSlot(int)
    def prog_count(self, count):
        self.main_ui.progressBar.setValue(count)

    @pyqtSlot()
    def thread_complete(self):
        QMessageBox.information(self, "Information", "Automation is completed.")
        self.main_ui.progressBar.setValue(0)

    @pyqtSlot(str)
    def thread_log(self, message):
        self.main_ui.tb_status.append(message)

    def exit_program(self):
        self.close()

    def init_setup(self):
        msg_type_list = ["REQUEST",
        "REQUEST_NO_RETURN",
        "REQUEST_ACK",
        "REQUEST_NO_RETURN_ACK",
        "NOTIFICATION",
        "RESPONSE",
        "RESPONSE_ACK",
        "ERROR",
        "ERROR_ACK",
        "TP_REQUEST",
        "TP_REQUEST_NO_RETURN",
        "TP_NOTIFICATION",
        "TP_RESPONSE",
        "TP_ERROR",
        "NOTIFICATION_ACK"]
        self.main_ui.cb_msg_type.addItems(msg_type_list)

        msg_status_list = ["E_OK",
        "E_NOT_OK",
        "E_UNKNOWN_SERVICE",
        "E_UNKNOWN_METHOD",
        "E_NOT_READY",
        "E_NOT_REACHABLE",
        "E_TIMEOUT",
        "E_WRONG_PROTOCOL_VERSION",
        "E_WRONG_INTERFACE_VERSION",
        "E_MALFORMED_MESSAGE",
        "E_WRONG_MESSAGE_TYPE"]
        self.main_ui.cb_msg_status.addItems(msg_status_list)

        self.main_ui.btn_run_automation.setEnabled(False)
        self.main_ui.btn_stop_automation.setEnabled(False)

class TransferThread(QThread):
    logSignal = pyqtSignal(str)
    countSignal = pyqtSignal(int)
    completeSignal = pyqtSignal()

    def __init__(self, excel_file_path, interval, cell_row_index):
        super(self.__class__, self).__init__()

        self.start_row = 8

        self.excel_file_path = excel_file_path
        wb = load_workbook(self.excel_file_path)
        print(wb.sheetnames)
        ws = wb.active

        self.total_count = ws['D2'].value

        self.test_data = []
        for row in ws.iter_rows(min_row=self.start_row, max_row=self.start_row + (self.total_count - 1)):
            row_data = [elem.value for elem in row[2:18]]
            self.test_data.append(row_data)
        wb.close()

        self.is_running = True
        self.interval = interval

    def run(self):
        count = 1
        wb = load_workbook(self.excel_file_path)
        ws = wb.active
        for data in self.test_data:
            address_sim = str(data[0])  # String
            port_sim = int(data[1])  # Int
            address_dest = str(data[2])  # String
            port_dest = int(data[3])  # Int
            protocol_ver = int(data[4])  # Int
            message_id = str(data[5])  # String
            service_id, method_id = getIds(message_id)
            msg_type = str(data[6])  # String
            msg_status = str(data[7])  # String
            client_id = int(data[8])  # Int
            session_id = int(data[9])  # Int
            payload = ConvertPayloadToHexString(str(data[10]))
            payload = hexStringToByte(payload)

            form = SomeIpForm(
                proto=protocol_ver,
                msg_type=msg_type,
                ret_code=msg_status,
                srv_id=service_id,
                method_id=method_id,
                client_id=client_id,
                session_id=session_id,
                payload=payload
            )

            form.setIp(address_sim, address_dest)
            form.setUdp(port_sim, port_dest)
            form.setPacket()
            # r, u = form.sr()

            # summary = r[0][1].summary()
            # original = r[0][1].original
            original = b'E\x00\x000\x1e\xe6\x00\x00@\x11\xb7\x1fd@n\x0bd@n,\xc7E\xc7E\x00\x1cFO\x00\x13\x05\x01\x00\x00\x00\x0c\x00\x01\x00\x01\x01\x01\x80\x00\x00\x00\x00\x00\x00\x00'
            summary = 'IP / UDP 100.64.110.11:51013 > 100.64.110.44:51013 / Raw / Padding'

            response = response_parser(original)
            ip_info = parse_ip(summary)
            response.update(ip_info)

            # response = {'from_ip':'100.64.111.44',
            #             'to_ip':'100.64.111.11',
            #             'service_id':'0x0013',
            #             'method_id':'0x0051',
            #             'message_type':'RESPONSE',
            #             'payload':'00000000'}
            ws[f'S{self.start_row + (count - 1)}']
            log_message = f'[{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}]\n'
            log_message += f'Source IP: {response["from_ip"]} / Dest IP: {response["to_ip"]}\n'
            log_message += f'Service ID: {response["service_id"]}\n'
            log_message += f'Method ID: {response["method_id"]}\n'
            log_message += f'Message Type: {response["message_type"]}\n'
            log_message += f'Payload: {response["payload"]}\n'

            self.logSignal.emit(log_message)

            self.countSignal.emit(count)
            count += 1

            if self.is_running is False:
                break

            time.sleep(self.interval)
        self.completeSignal.emit()
        wb.save(self.excel_file_path)

    def stop_thread(self):
        self.is_running = False


if __name__ == "__main__":
    app = QApplication(sys.argv)
    myWindow = MainDialog()
    myWindow.show()
    app.exec()